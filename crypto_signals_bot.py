import ccxt.async_support as ccxt
import pandas as pd
from ta.trend import IchimokuIndicator, MACD, ADXIndicator, PSARIndicator, VortexIndicator, CCIIndicator, EMAIndicator
from ta.volatility import BollingerBands, AverageTrueRange, KeltnerChannel, DonchianChannel
from ta.momentum import RSIIndicator, StochasticOscillator, WilliamsRIndicator, ROCIndicator
from ta.volume import OnBalanceVolumeIndicator, ChaikinMoneyFlowIndicator
import asyncio
import numpy as np
from datetime import datetime, timedelta
import logging
import os
import pickle
import json
from telegram import Bot
import aiofiles
import sys
import csv
import ta
import subprocess
import aiohttp
import openpyxl
from openpyxl.utils import get_column_letter
from scipy.stats import pearsonr
from scipy.signal import argrelextrema
import argparse
# Import the ML module
from ml_models import CryptoMLPredictor, batch_train_models, ohlcv_to_dataframe, ModelType
# Import pattern recognition module
from pattern_recognition import CandlestickPatterns, ChartPatterns, analyze_patterns
import joblib

# Настройка SelectorEventLoop для Windows
if sys.platform == 'win32':
    asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())

# Настройка логирования
logging.basicConfig(
    format='%(asctime)s [%(levelname)s] %(message)s',
    level=logging.INFO,
    handlers=[
        logging.FileHandler('signals.log'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# Отключение логов для ccxt и связанных библиотек
for name in logging.root.manager.loggerDict:
    if name.startswith('ccxt') or name in ['httpx', 'httpcore', 'urllib3', 'websockets', 'asyncio']:
        logging.getLogger(name).setLevel(logging.ERROR)

# Проверка версии ta
try:
    ta_version = ta.__version__
except AttributeError:
    try:
        ta_version = subprocess.check_output([sys.executable, '-m', 'pip', 'show', 'ta']).decode()
        ta_version = next((line.split(': ')[1] for line in ta_version.split('\n') if line.startswith('Version: ')), 'unknown')
    except Exception:
        ta_version = 'unknown'
logger.info(f"Используется версия библиотеки ta: {ta_version}")
if ta_version < '0.11.0' and ta_version != 'unknown':
    logger.warning(f"Версия библиотеки ta ({ta_version}) устарела. Рекомендуется обновить до >=0.11.0: pip install ta --upgrade")

# Загрузка конфигурации
CONFIG_FILE = 'config.json'
DEFAULT_CONFIG = {
    'telegram_token': '8143467986:AAEo0IC1jwEpHXeNNIw1MmQORG-JPjU65nE',
    'telegram_chat_id': '-1002663850031',
    'cryptopanic_api_key': 'ac5c98562407cf9d569c4c453b02f5b2b50b79ae',
    'timeframe': '1h',
    'min_volume_usd': 1_000_000,
    'signal_threshold': 0.52,  # 12/23 индикаторов
    'pump_price_threshold': 10.0,  # % роста за час
    'pump_volume_multiplier': 3.0,  # Volume Spike
    'tp_percent': 5.0,  # %
    'sl_percent': 2.0,  # %
    'leverage': 10,
    'meme_coins': [
        "DOGE", "SHIB", "PEPE", "FLOKI", "WIF", "BONK", "BRETT", "BOME",
        "MEW", "MOG", "POPCAT", "QUBIC", "PEIPEI", "TURBO", "SUN", "NEIRO",
        "GME", "AMC", "SPX", "ZRO", "NOT", "RATS", "DOGS", "HMSTR"
    ],
    'meme_leverage': 5,
    'meme_tp_percent': 3.0,  # %
    'atr_volatility_multiplier': 1.0,  # Для фильтра волатильности
    'signal_expiry_hours': 12,
    'position_size_usd': 1000,
    'ohlcv_limit': 60,  # Уменьшено для поддержки новых пар
    # ML parameters
    'use_ml_predictions': True,
    'ml_confidence_threshold': 0.65,  # Минимальная достоверность для учёта ML-прогноза
    'ml_retrain_days': 7,  # Переобучение моделей каждые 7 дней
    'min_ml_accuracy': 0.60,  # Минимальная точность ML-модели для использования прогнозов
    # Pattern recognition parameters
    'use_pattern_recognition': True,
    'pattern_lookback_period': 100,  # Number of candles to look back for chart patterns
    'pattern_weight': 2.0,  # Weight for pattern signals (candlestick patterns + chart patterns)
    # Advanced ML settings
    'default_model_type': ModelType.RANDOM_FOREST  # Default ML model type to use
}

def load_config():
    try:
        with open(CONFIG_FILE, 'r') as f:
            return json.load(f)
    except FileNotFoundError:
        with open(CONFIG_FILE, 'w') as f:
            json.dump(DEFAULT_CONFIG, f, indent=4)
        return DEFAULT_CONFIG

# Функция для преобразования строкового представления ModelType в объект ModelType
def convert_model_type(model_type_str):
    if isinstance(model_type_str, str):
        if model_type_str.lower() == "random_forest":
            return ModelType.RANDOM_FOREST
        elif model_type_str.lower() == "gradient_boosting":
            return ModelType.GRADIENT_BOOSTING
        elif model_type_str.lower() == "ensemble":
            return ModelType.ENSEMBLE
        elif model_type_str.lower() == "lstm":
            return ModelType.LSTM
    return ModelType.RANDOM_FOREST  # По умолчанию, если не распознано

config = load_config()

# Преобразуем model_type из строки в объект ModelType, если необходимо
if 'default_model_type' in config and isinstance(config['default_model_type'], str):
    config['default_model_type'] = convert_model_type(config['default_model_type'])

# Настройка Telegram и CryptoPanic
TELEGRAM_TOKEN = config['telegram_token']
TELEGRAM_CHAT_ID = config['telegram_chat_id']
CRYPTOPANIC_API_KEY = config.get('cryptopanic_api_key', '')

# Настройка Bybit
bybit = ccxt.bybit({
    'enableRateLimit': True,
    'defaultType': 'swap',
    'rateLimit': 100
})

# Тест Telegram при старте
async def test_telegram():
    try:
        bot = Bot(token=TELEGRAM_TOKEN)
        await bot.send_message(chat_id=TELEGRAM_CHAT_ID, text="🚀 Бот запущен! Ожидаю сигналы...")
        logger.info("Тестовое сообщение отправлено в Telegram")
    except Exception as e:
        logger.error(f"Ошибка тестового сообщения в Telegram: {e}")

# Очистка старого кэша
def clean_old_cache(cache_dir='cache', max_age_hours=24):
    os.makedirs(cache_dir, exist_ok=True)
    now = datetime.now()
    for filename in os.listdir(cache_dir):
        file_path = os.path.join(cache_dir, filename)
        if os.path.isfile(file_path):
            file_mtime = datetime.fromtimestamp(os.path.getmtime(file_path))
            if (now - file_mtime).total_seconds() > max_age_hours * 3600:
                os.remove(file_path)
                logger.debug(f"Удалён старый кэш: {file_path}")

# Загрузка рынков
async def load_markets():
    try:
        markets = await bybit.load_markets()
        symbols = [symbol for symbol, info in markets.items() if info['type'] == 'swap' and info['active'] and 
                  (symbol.endswith('/USDT:USDT') or symbol.endswith('/USDT')) and not 'BUSD' in symbol]
        
        # Преобразуем Bybit символы к стандартному формату
        standardized_symbols = [
            symbol.replace(':USDT', '') for symbol in symbols
        ]
        
        logger.info(f"Загружено {len(standardized_symbols)} активных USDT фьючерсных пар на Bybit")
        return standardized_symbols
    except Exception as e:
        await send_telegram_message(f"⚠️ Ошибка загрузки рынков Bybit: {e}")
        logger.error(f"Ошибка загрузки рынков Bybit: {e}")
        return []

# Кэширование OHLCV
async def load_ohlcv(symbol, timeframe=config['timeframe'], limit=config['ohlcv_limit']):
    cache_file = f'cache/{symbol.replace("/", "_")}_{timeframe}.pkl'
    os.makedirs('cache', exist_ok=True)
    
    try:
        if os.path.exists(cache_file):
            file_age = datetime.now() - datetime.fromtimestamp(os.path.getmtime(cache_file))
            if file_age.total_seconds() < 3600:
                async with aiofiles.open(cache_file, 'rb') as f:
                    ohlcv = pickle.loads(await f.read())
                    logger.debug(f"Загружен актуальный кэш для {symbol}")
                    return ohlcv

        ohlcv = await bybit.fetch_ohlcv(symbol, timeframe, limit=limit)
        async with aiofiles.open(cache_file, 'wb') as f:
            await f.write(pickle.dumps(ohlcv))
        logger.debug(f"Обновлён кэш для {symbol}")
        return ohlcv
    except Exception as e:
        logger.error(f"Ошибка загрузки OHLCV для {symbol}: {e}")
        if os.path.exists(cache_file):
            async with aiofiles.open(cache_file, 'rb') as f:
                ohlcv = pickle.loads(await f.read())
                logger.warning(f"Используем старый кэш для {symbol} из-за ошибки")
                return ohlcv
        return None

# Получение Fear & Greed Index
async def fetch_fear_and_greed():
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get('https://api.alternative.me/fng/') as response:
                data = await response.json()
                fng_value = int(data['data'][0]['value'])
                logger.debug(f"Fear & Greed Index: {fng_value}")
                return fng_value
    except Exception as e:
        logger.error(f"Ошибка получения Fear & Greed Index: {e}")
        return 50  # Fallback значение

# Проверка новостей через CryptoPanic
async def check_news(symbol, price_change_1h):
    if not CRYPTOPANIC_API_KEY or abs(price_change_1h) < 10:
        return "Не проверялись"
    
    try:
        base_currency = symbol.split('/')[0].replace('1000', '').replace('10000', '').replace('1000000', '').replace('10000000', '')
        url = f"https://cryptopanic.com/api/v1/posts/?auth_token={CRYPTOPANIC_API_KEY}&currencies={base_currency}&filter=important,hot"
        async with aiohttp.ClientSession() as session:
            async with session.get(url) as response:
                data = await response.json()
                if not data.get('results'):
                    return "Не найдены"
                
                # Проверяем новости за последние 24 часа
                recent_news = [
                    news for news in data['results']
                    if datetime.fromisoformat(news['published_at'].replace('Z', '+00:00')) > datetime.utcnow() - timedelta(hours=24)
                ]
                
                if not recent_news:
                    return "Не найдены"
                
                # Определяем настроение первой новости
                sentiment = recent_news[0].get('metadata', {}).get('sentiment', 'neutral')
                sentiment_str = "Позитивные" if sentiment == "positive" else "Негативные" if sentiment == "negative" else "Нейтральные"
                return f"{sentiment_str} (Δцена {price_change_1h:+.1f}%)"
    except Exception as e:
        logger.error(f"Ошибка проверки новостей для {symbol}: {e}")
        return "Не проверялись"

# Фильтр по объёму торгов с кэшированием
async def filter_high_volume_symbols(symbols, min_volume_usd=config['min_volume_usd']):
    cache_file = 'cache/high_volume_symbols.pkl'
    cache_max_age = 3600

    if os.path.exists(cache_file):
        file_age = datetime.now() - datetime.fromtimestamp(os.path.getmtime(cache_file))
        if file_age.total_seconds() < cache_max_age:
            async with aiofiles.open(cache_file, 'rb') as f:
                cached_data = pickle.loads(await f.read())
                logger.info(f"Загружен кэш фильтра объёма: {len(cached_data)} пар")
                return cached_data, {}

    high_volume_symbols = []
    volume_data = {}
    logger.info(f"Фильтрация {len(symbols)} пар по объёму >${min_volume_usd/1_000_000:.2f}M")
    
    async def check_volume(symbol):
        try:
            ticker = await bybit.fetch_ticker(symbol)
            volume_usd = ticker['baseVolume'] * ticker['close']
            
            # Подробное логирование объемов
            logger.info(f"Объём для {symbol}: ${volume_usd/1_000_000:.2f}M")
            
            if volume_usd > min_volume_usd:
                logger.debug(f"Пара {symbol}: Объём ${volume_usd/1_000_000:.2f}M (прошла фильтр)")
                return symbol, volume_usd
            logger.debug(f"Пара {symbol}: Объём ${volume_usd/1_000_000:.2f}M (отфильтрована)")
            return None, None
        except Exception as e:
            logger.error(f"Ошибка получения объёма для {symbol}: {e}")
            return None, None

    chunk_size = 5
    for i in range(0, len(symbols), chunk_size):
        chunk = symbols[i:i + chunk_size]
        results = await asyncio.gather(*[check_volume(symbol) for symbol in chunk])
        for symbol, volume_usd in results:
            if symbol:
                high_volume_symbols.append(symbol)
                volume_data[symbol] = volume_usd
        await asyncio.sleep(1)

    async with aiofiles.open(cache_file, 'wb') as f:
        await f.write(pickle.dumps(high_volume_symbols))
    
    logger.info(f"Фильтрация завершена: {len(high_volume_symbols)} пар прошли фильтр объёма")
    return high_volume_symbols, volume_data

# Расчёт VWAP
def calculate_vwap(df, sigma=1.0):
    try:
        typical_price = (df['high'] + df['low'] + df['close']) / 3
        vwap = (typical_price * df['volume']).cumsum() / df['volume'].cumsum()
        vwap_dev = typical_price.rolling(window=20).std()
        return vwap, vwap_dev * sigma
    except Exception as e:
        logger.error(f"Ошибка расчёта VWAP: {e}")

# Расчёт уровней Фибоначчи
def calculate_fibonacci_levels(df, period=20):
    try:
        high = df['high'].rolling(window=period).max()
        low = df['low'].rolling(window=period).min()
        diff = high - low
        fib_382 = low + diff * 0.382
        fib_618 = low + diff * 0.618
        return fib_382, fib_618
    except Exception as e:
        logger.error(f"Ошибка расчёта уровней Фибоначчи: {e}")
        return pd.Series(np.nan, index=df.index), pd.Series(np.nan, index=df.index)

# Экспорт CSV в Excel
def export_csv_to_excel(csv_file='signals_history.csv', excel_file='signals_history.xlsx'):
    try:
        # Чтение CSV с корректной кодировкой
        df = pd.read_csv(csv_file, encoding='utf-8-sig')
        df.to_excel(excel_file, index=False, engine='openpyxl')
        
        # Автоматическая настройка ширины столбцов
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        wb.save(excel_file)
        
        logger.debug(f"Экспортировано в Excel: {excel_file}")
    except Exception as e:
        logger.error(f"Ошибка экспорта в Excel: {e}")
        # Попытка с другой кодировкой при ошибке
        try:
            df = pd.read_csv(csv_file, encoding='latin1')
            df.to_excel(excel_file, index=False, engine='openpyxl')
            logger.debug(f"Экспортировано в Excel с кодировкой latin1: {excel_file}")
        except Exception as e2:
            logger.error(f"Повторная ошибка экспорта в Excel: {e2}")

# Аналитика по исходам
def generate_analytics():
    try:
        df = pd.read_csv('signals_history.csv')
        if df.empty:
            return "Аналитика: Нет данных о сигналах."

        # Парсинг индикаторов
        df['indicators'] = df['indicators'].apply(lambda x: eval(x) if isinstance(x, str) else x)

        # Общая статистика
        total_signals = len(df)
        tp_count = len(df[df['outcome'] == 'TP'])
        sl_count = len(df[df['outcome'] == 'SL'])
        expired_count = len(df[df['outcome'] == 'Expired'])
        win_rate = (tp_count / total_signals * 100) if total_signals > 0 else 0
        avg_pl_percent = df['profit_loss_percent'].mean()
        avg_pl_usd = df['profit_loss_usd'].mean()

        analytics = f"📊 Аналитика сигналов ({total_signals}):\n"
        analytics += f"  TP: {tp_count} ({tp_count/total_signals*100:.1f}%)\n"
        analytics += f"  SL: {sl_count} ({sl_count/total_signals*100:.1f}%)\n"
        analytics += f"  Expired: {expired_count} ({expired_count/total_signals*100:.1f}%)\n"
        analytics += f"  Общий винрейт: {win_rate:.1f}%\n"
        analytics += f"  Средний P/L: {avg_pl_percent:+.2f}% (${avg_pl_usd:+.2f})\n"

        # Статистика по парам
        analytics += "\n🔍 По парам:\n"
        for symbol in df['symbol'].unique():
            symbol_df = df[df['symbol'] == symbol]
            symbol_signals = len(symbol_df)
            symbol_tp = len(symbol_df[symbol_df['outcome'] == 'TP'])
            symbol_win_rate = (symbol_tp / symbol_signals * 100) if symbol_signals > 0 else 0
            symbol_avg_pl = symbol_df['profit_loss_percent'].mean()
            analytics += f"  {symbol}: {symbol_signals} сигналов, {symbol_win_rate:.1f}% TP, средний P/L {symbol_avg_pl:+.2f}%\n"

        # Статистика по типам сигналов
        analytics += "\n📈 По типам:\n"
        for signal_type in ['long', 'short']:
            type_df = df[df['type'] == signal_type]
            type_signals = len(type_df)
            type_tp = len(type_df[type_df['outcome'] == 'TP'])
            type_win_rate = (type_tp / type_signals * 100) if type_signals > 0 else 0
            type_avg_pl = type_df['profit_loss_percent'].mean()
            analytics += f"  {signal_type.capitalize()}: {type_signals} сигналов, {type_win_rate:.1f}% TP, средний P/L {type_avg_pl:+.2f}%\n"

        # Винрейт по индикаторам
        analytics += "\n🔧 Винрейт по индикаторам:\n"
        indicator_names = {
            'price_above_cloud': 'Ichimoku Cloud',
            'price_below_cloud': 'Ichimoku Cloud',
            'macd_above_signal': 'MACD Above Signal',
            'macd_below_signal': 'MACD Below Signal',
            'rsi_oversold': 'RSI Oversold',
            'rsi_overbought': 'RSI Overbought',
            'price_below_bb_lower': 'Bollinger Bands Lower',
            'price_above_bb_upper': 'Bollinger Bands Upper',
            'ema_crossover_long': 'EMA Crossover',
            'ema_crossover_short': 'EMA Crossover',
            'stoch_oversold': 'Stochastic Oversold',
            'stoch_overbought': 'Stochastic Overbought',
            'obv_uptrend': 'OBV Uptrend',
            'obv_downtrend': 'OBV Downtrend',
            'parabolic_sar_long': 'Parabolic SAR',
            'parabolic_sar_short': 'Parabolic SAR',
            'adx_strong_trend': 'ADX Strong Trend',
            'volume_spike': 'Volume Spike',
            'rsi_divergence_bullish': 'RSI Bullish Divergence',
            'rsi_divergence_bearish': 'RSI Bearish Divergence',
            'price_above_vwap': 'Price Above VWAP',
            'price_below_vwap': 'Price Below VWAP',
            'cci_oversold': 'CCI Oversold',
            'cci_overbought': 'CCI Overbought',
            'williams_r_oversold': 'Williams %R Oversold',
            'williams_r_overbought': 'Williams %R Overbought',
            'fibonacci_382_long': 'Fibonacci 38.2% Long',
            'fibonacci_382_short': 'Fibonacci 38.2% Short',
            'vortex_long': 'Vortex Long',
            'vortex_short': 'Vortex Short',
            'cmf_long': 'Chaikin Money Flow Long',
            'cmf_short': 'Chaikin Money Flow Short',
            'keltner_long': 'Keltner Channel Long',
            'keltner_short': 'Keltner Channel Short',
            'roc_long': 'Rate of Change Long',
            'roc_short': 'Rate of Change Short',
            'donchian_long': 'Donchian Channel Long',
            'donchian_short': 'Donchian Channel Long',
            'fng_fear': 'Fear & Greed Fear',
            'fng_greed': 'Fear & Greed Greed'
        }
        indicator_stats = {}
        for ind in df['indicators'].iloc[0].keys():
            ind_signals = len(df[df['indicators'].apply(lambda x: x.get(ind))])
            ind_tp = len(df[(df['indicators'].apply(lambda x: x.get(ind))) & (df['outcome'] == 'TP')])
            ind_win_rate = (ind_tp / ind_signals * 100) if ind_signals > 0 else 0
            if ind_signals >= 5:  # Минимум 5 сигналов для статистики
                indicator_stats[ind] = (ind_win_rate, ind_signals)
        sorted_indicators = sorted(indicator_stats.items(), key=lambda x: x[1][0], reverse=True)
        for ind, (win_rate, signals) in sorted_indicators[:10]:  # Топ-10 индикаторов
            analytics += f"  {indicator_names.get(ind, ind)}: {win_rate:.1f}% ({signals} сигналов)\n"

        # Корреляция индикаторов с P/L
        analytics += "\n📉 Корреляция с P/L:\n"
        if len(df) >= 10:  # Минимум 10 сигналов для корреляции
            correlations = {}
            for ind in df['indicators'].iloc[0].keys():
                ind_values = df['indicators'].apply(lambda x: 1 if x.get(ind) else 0)
                if ind_values.sum() > 0:  # Индикатор должен быть активен хотя бы раз
                    corr, _ = pearsonr(ind_values, df['profit_loss_percent'])
                    if not np.isnan(corr):
                        correlations[ind] = corr
            sorted_correlations = sorted(correlations.items(), key=lambda x: abs(x[1]), reverse=True)
            for ind, corr in sorted_correlations[:5]:  # Топ-5 корреляций
                analytics += f"  {indicator_names.get(ind, ind)}: {corr:+.2f}\n"
        else:
            analytics += "  Недостаточно данных для корреляции.\n"

        # Запись в файл
        with open('analytics_report.txt', 'w') as f:
            f.write(analytics)

        # Экспорт в Excel
        analytics_df = pd.DataFrame({
            'Metric': [],
            'Value': []
        })
        for line in analytics.split('\n'):
            if ':' in line:
                metric, value = line.split(':', 1)
                analytics_df = pd.concat([analytics_df, pd.DataFrame({'Metric': [metric.strip()], 'Value': [value.strip()]})], ignore_index=True)
        analytics_df.to_excel('analytics_report.xlsx', index=False, engine='openpyxl')
        
        # Настройка ширины столбцов
        wb = openpyxl.load_workbook('analytics_report.xlsx')
        ws = wb.active
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        wb.save('analytics_report.xlsx')
        
        logger.info("Аналитика сгенерирована и экспортирована")
        return analytics
    except Exception as e:
        logger.error(f"Ошибка генерации аналитики: {e}")
        return f"Ошибка аналитики: {e}"

# Отправка сообщения в Telegram
async def send_telegram_message(message):
    try:
        bot = Bot(token=TELEGRAM_TOKEN)
        await bot.send_message(chat_id=TELEGRAM_CHAT_ID, text=message)
        logger.info(f"Отправлено сообщение в Telegram: {message[:100]}...")
    except Exception as e:
        logger.error(f"Ошибка отправки в Telegram: {e}")

# Запись сигналов в CSV
def log_signal_to_csv(signal):
    """Запись сигнала в CSV-файл"""
    file_exists = os.path.isfile('signals_history.csv')
    
    # Подготовка данных для CSV
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    symbol = signal['symbol']
    signal_type = signal['type']
    entry_price = signal['entry_price']
    tp = signal['tp']
    sl = signal['sl']
    leverage = signal['leverage']
    
    # ML prediction data
    ml_prediction = signal.get('ml_prediction')
    ml_direction = ml_prediction['direction'] if ml_prediction else 'none'
    ml_confidence = ml_prediction['confidence'] if ml_prediction else 0
    
    # Indicators summary
    total_indicators = len(signal['indicators'])
    positive_indicators = sum(1 for v in signal['indicators'].values() if v)
    
    # High confidence combos
    high_confidence = ','.join(signal['high_confidence_combo']) if signal['high_confidence_combo'] else 'none'
    
    # CSV fields
    fields = [
        now,                   # Timestamp
        symbol,                # Symbol
        signal_type,           # Signal type (long/short)
        entry_price,           # Entry price
        tp,                    # Take profit
        sl,                    # Stop loss
        leverage,              # Leverage
        signal['volume_usd'],  # Volume USD
        signal['news'],        # News context
        'active',              # Status (active/tp/sl/expired)
        None,                  # Exit price (filled when closed)
        None,                  # P/L % (filled when closed)
        None,                  # P/L $ (filled when closed)
        None,                  # Exit time (filled when closed)
        positive_indicators,   # Number of positive indicators
        total_indicators,      # Total indicators
        high_confidence,       # High confidence combination
        ml_direction,          # ML prediction direction
        ml_confidence          # ML prediction confidence
    ]
    
    try:
        with open('signals_history.csv', 'a', newline='') as f:
            writer = csv.writer(f)
            
            # Write header if file doesn't exist
            if not file_exists:
                writer.writerow([
                    'timestamp', 'symbol', 'type', 'entry_price', 'tp', 'sl', 'leverage',
                    'volume_usd', 'news', 'status', 'exit_price', 'pl_percent', 'pl_usd', 'exit_time',
                    'positive_indicators', 'total_indicators', 'high_confidence', 
                    'ml_direction', 'ml_confidence'
                ])
            
            writer.writerow(fields)
            logger.debug(f"Сигнал {symbol} {signal_type} записан в CSV")
    except Exception as e:
        logger.error(f"Ошибка записи сигнала в CSV: {e}")

    # Обновление Excel после добавления строки
    asyncio.create_task(update_excel())

# Обновление сигнала в CSV
def update_signal_in_csv(signal):
    csv_file = 'signals_history.csv'
    temp_file = 'signals_history_temp.csv'
    found = False

    with open(csv_file, 'r') as f, open(temp_file, 'w', newline='') as temp:
        reader = csv.reader(f)
        writer = csv.writer(temp)
        header = next(reader)
        writer.writerow(header)

        for row in reader:
            if (row[0] == signal['timestamp'] and
                row[1] == signal['symbol'] and
                row[2] == signal['type']):
                writer.writerow([
                    signal['timestamp'],
                    signal['symbol'],
                    signal['type'],
                    signal['entry'],
                    signal['tp'],
                    signal['sl'],
                    signal['leverage'],
                    str(signal['indicators']),
                    signal.get('exit_price', ''),
                    signal.get('outcome', ''),
                    signal.get('profit_loss_percent', ''),
                    signal.get('profit_loss_usd', ''),
                    signal.get('exit_timestamp', '')
                ])
                found = True
            else:
                writer.writerow(row)

    if found:
        os.replace(temp_file, csv_file)
        logger.debug(f"Обновлена запись в CSV для {signal['symbol']}")
        export_csv_to_excel()
    else:
        logger.warning(f"Сигнал {signal['symbol']} не найден в CSV для обновления")
        os.remove(temp_file)

# Уведомление об исходе
async def send_outcome_notification(signal):
    try:
        signal_type = "Покупка" if signal['type'] == 'long' else "Продажа"
        message = f"📊 Результат сигнала: {signal_type} {signal['symbol']}\n"
        message += f"💰 Вход: ${signal['entry']:.4f}\n"
        message += f"📉 Выход: ${signal['exit_price']:.4f}\n"
        message += f"🔔 Исход: {signal['outcome']}\n"
        message += f"📈 P/L: {signal['profit_loss_percent']:+.2f}% (${signal['profit_loss_usd']:+.2f})\n"
        message += f"🕒 Время закрытия: {signal['exit_timestamp']}\n"
        message += f"⚠️ Не финансовая рекомендация!"
        
        await send_telegram_message(message)
        logger.info(f"Уведомление об исходе отправлено: {signal['type']} на {signal['symbol']}")
    except Exception as e:
        logger.error(f"Ошибка отправки уведомления об исходе: {e}")

# Мониторинг исхода сигнала
async def monitor_signal_outcome(signal, expiry_hours=config['signal_expiry_hours']):
    try:
        symbol = signal['symbol']
        signal_type = signal['type']
        entry_price = signal['entry']
        tp = signal['tp']
        sl = signal['sl']
        leverage = signal['leverage']
        position_size = config['position_size_usd']
        start_time = pd.to_datetime(signal['timestamp'])
        expiry_time = start_time + timedelta(hours=expiry_hours)

        logger.debug(f"Начало мониторинга сигнала: {signal_type} на {symbol}")

        while datetime.utcnow() < expiry_time:
            try:
                ticker = await bybit.fetch_ticker(symbol)
                current_price = ticker['last']
                logger.debug(f"{symbol}: Текущая цена ${current_price:.4f}")

                # Проверка TP/SL
                if signal_type == 'long':
                    if current_price >= tp:
                        outcome = 'TP'
                        exit_price = tp
                        profit_loss_percent = (tp - entry_price) / entry_price * 100 * leverage
                        profit_loss_usd = position_size * (tp / entry_price - 1) * leverage
                        break
                    elif current_price <= sl:
                        outcome = 'SL'
                        exit_price = sl
                        profit_loss_percent = (sl - entry_price) / entry_price * 100 * leverage
                        profit_loss_usd = position_size * (sl / entry_price - leverage)
                        break
                else:  # short
                    if current_price <= tp:
                        outcome = 'TP'
                        exit_price = tp
                        profit_loss_percent = (entry_price - tp) / entry_price * 100 * leverage
                        profit_loss_usd = position_size * (entry_price / tp - 1) * leverage
                        break
                    elif current_price >= sl:
                        outcome = 'SL'
                        exit_price = sl
                        profit_loss_percent = (entry_price - sl) / entry_price * 100 * leverage
                        profit_loss_usd = position_size * (entry_price / sl - 1) * leverage
                        break

                await asyncio.sleep(7200)  # Проверка каждые 2 часа
            except Exception as e:
                logger.error(f"Ошибка мониторинга {symbol}: {e}")
                await asyncio.sleep(300)  # Пауза 5 минут при ошибке

        else:
            # Сигнал истёк
            outcome = 'Expired'
            exit_price = current_price
            profit_loss_percent = 0
            profit_loss_usd = 0

        exit_timestamp = datetime.utcnow().isoformat() + 'Z'
        signal.update({
            'exit_price': exit_price,
            'outcome': outcome,
            'profit_loss_percent': profit_loss_percent,
            'profit_loss_usd': profit_loss_usd,
            'exit_timestamp': exit_timestamp
        })

        logger.info(f"Сигнал {signal_type} на {symbol}: Исход = {outcome}, P/L = {profit_loss_percent:.2f}% (${profit_loss_usd:.2f})")

        # Обновление CSV и экспорт в Excel
        update_signal_in_csv(signal)

        # Отправка уведомления об исходе
        await send_outcome_notification(signal)

        return signal
    except Exception as e:
        logger.error(f"Ошибка мониторинга исхода для {symbol}: {e}")
        return None

# Расчёт индикаторов с обработкой ошибок
def calculate_indicators(df, fng_value):
    """Расчёт технических индикаторов для анализа"""
    try:
        if len(df) < 14:
            return None
            
        df = df.copy()
        
        # Расчёт основных индикаторов
        current_price = df['close'].iloc[-1]
        
        # RSI
        rsi = RSIIndicator(close=df['close'], window=14)
        df['rsi'] = rsi.rsi()
        
        # MACD
        macd = MACD(close=df['close'])
        df['macd'] = macd.macd()
        df['macd_signal'] = macd.macd_signal()
        df['macd_diff'] = macd.macd_diff()
        
        # Bollinger Bands
        bb = BollingerBands(close=df['close'], window=20, window_dev=2)
        df['bb_upper'] = bb.bollinger_hband()
        df['bb_lower'] = bb.bollinger_lband()
        df['bb_middle'] = bb.bollinger_mavg()
        df['bb_width'] = (df['bb_upper'] - df['bb_lower']) / df['bb_middle']
        
        # ATR (используется для фильтрации по волатильности)
        atr = AverageTrueRange(high=df['high'], low=df['low'], close=df['close'], window=14)
        df['atr'] = atr.average_true_range()
        df['avg_atr'] = df['atr'].rolling(window=14).mean()
        
        # EMA
        df['ema_9'] = ta.trend.ema_indicator(df['close'], window=9)
        df['ema_21'] = ta.trend.ema_indicator(df['close'], window=21)
        df['ema_50'] = ta.trend.ema_indicator(df['close'], window=50)
        df['ema_200'] = ta.trend.ema_indicator(df['close'], window=200)
        
        # Stochastic
        stoch = StochasticOscillator(high=df['high'], low=df['low'], close=df['close'])
        df['stoch_k'] = stoch.stoch()
        df['stoch_d'] = stoch.stoch_signal()
        
        # ADX
        adx = ADXIndicator(high=df['high'], low=df['low'], close=df['close'])
        df['adx'] = adx.adx()
        df['adx_plus_di'] = adx.adx_pos()
        df['adx_minus_di'] = adx.adx_neg()
        
        # OBV
        obv = OnBalanceVolumeIndicator(close=df['close'], volume=df['volume'])
        df['obv'] = obv.on_balance_volume()
        df['obv_ema'] = ta.trend.ema_indicator(df['obv'], window=21)
        
        # Ichimoku
        ichimoku = IchimokuIndicator(high=df['high'], low=df['low'])
        df['tenkan_sen'] = ichimoku.ichimoku_conversion_line()
        df['kijun_sen'] = ichimoku.ichimoku_base_line()
        df['senkou_span_a'] = ichimoku.ichimoku_a()
        df['senkou_span_b'] = ichimoku.ichimoku_b()
        
        # Parabolic SAR
        psar = PSARIndicator(high=df['high'], low=df['low'], close=df['close'])
        df['psar'] = psar.psar()
        
        # Williams %R
        wr = WilliamsRIndicator(high=df['high'], low=df['low'], close=df['close'])
        df['wr'] = wr.williams_r()
        
        # Supertrend
        atr_period = config.get('supertrend_atr_period', 14)
        multiplier = config.get('supertrend_multiplier', 3.0)
        
        # Используем расчитанный ранее ATR для суперттренда
        df['basic_upper'] = (df['high'] + df['low']) / 2 + multiplier * df['atr']
        df['basic_lower'] = (df['high'] + df['low']) / 2 - multiplier * df['atr']
        
        # Инициализируем столбцы для суперттренда
        df['supertrend'] = 0.0
        df['supertrend_direction'] = 0  # 1 для бычьего тренда, -1 для медвежьего
        
        # Первое значение суперттренда
        if len(df) > atr_period:
            # Устанавливаем начальное значение
            df.loc[atr_period, 'supertrend'] = df.loc[atr_period, 'basic_upper']
            df.loc[atr_period, 'supertrend_direction'] = -1  # Начинаем с медвежьего тренда
            
            # Рассчитываем суперттренд для каждой свечи
            for i in range(atr_period + 1, len(df)):
                prev_supertrend = df.loc[i-1, 'supertrend']
                prev_direction = df.loc[i-1, 'supertrend_direction']
                curr_close = df.loc[i, 'close']
                curr_upper = df.loc[i, 'basic_upper']
                curr_lower = df.loc[i, 'basic_lower']
                
                # Логика расчета суперттренда
                if prev_supertrend == prev_supertrend:  # Проверка на NaN
                    if prev_supertrend <= df.loc[i-1, 'close'] and prev_direction <= 0:
                        # Переход к бычьему тренду
                        df.loc[i, 'supertrend'] = curr_lower
                        df.loc[i, 'supertrend_direction'] = 1
                    elif prev_supertrend > df.loc[i-1, 'close'] and prev_direction >= 0:
                        # Переход к медвежьему тренду
                        df.loc[i, 'supertrend'] = curr_upper
                        df.loc[i, 'supertrend_direction'] = -1
                    elif prev_direction == 1:
                        # Продолжение бычьего тренда
                        df.loc[i, 'supertrend'] = max(curr_lower, prev_supertrend)
                        df.loc[i, 'supertrend_direction'] = 1
                    else:
                        # Продолжение медвежьего тренда
                        df.loc[i, 'supertrend'] = min(curr_upper, prev_supertrend)
                        df.loc[i, 'supertrend_direction'] = -1
                else:
                    # Если предыдущее значение NaN, используем базовые значения
                    df.loc[i, 'supertrend'] = curr_upper
                    df.loc[i, 'supertrend_direction'] = -1
            
            # Логирование расчета Суперттренда
            supertrend_value = df['supertrend'].iloc[-1]
            supertrend_direction = df['supertrend_direction'].iloc[-1]
            current_close = df['close'].iloc[-1]
            
            supertrend_long = current_close > supertrend_value
            supertrend_short = current_close < supertrend_value
            
            logger.debug(f"Supertrend: значение={supertrend_value:.6f}, направление={'UP' if supertrend_direction > 0 else 'DOWN'}")
            logger.debug(f"Supertrend сигналы: Long={supertrend_long}, Short={supertrend_short}, текущая цена={current_close:.6f}")
            
            if supertrend_long:
                logger.debug(f"Цена выше Supertrend на {((current_close - supertrend_value) / supertrend_value * 100):.2f}%")
            elif supertrend_short:
                logger.debug(f"Цена ниже Supertrend на {((supertrend_value - current_close) / supertrend_value * 100):.2f}%")
        
        # Проверка пересечений и условий
        indicators = {
            'current_price': current_price,
            'rsi': df['rsi'].iloc[-1],
            'macd': df['macd'].iloc[-1],
            'macd_signal': df['macd_signal'].iloc[-1],
            'bb_width': df['bb_width'].iloc[-1],
            'atr': df['atr'].iloc[-1],
            'avg_atr': df['avg_atr'].iloc[-1],
            'fng_value': fng_value,
            
            # Условия для Long
            'price_above_cloud': df['close'].iloc[-1] > df['senkou_span_a'].iloc[-1] and df['close'].iloc[-1] > df['senkou_span_b'].iloc[-1],
            'macd_above_signal': df['macd'].iloc[-1] > df['macd_signal'].iloc[-1],
            'rsi_oversold': df['rsi'].iloc[-1] < 40 and df['rsi'].iloc[-1] > df['rsi'].iloc[-2],
            'price_below_bb_lower': df['close'].iloc[-1] < df['bb_lower'].iloc[-1],
            'ema_crossover_long': df['ema_9'].iloc[-1] > df['ema_21'].iloc[-1] and df['ema_9'].iloc[-2] <= df['ema_21'].iloc[-2],
            'stoch_oversold': df['stoch_k'].iloc[-1] < 20 and df['stoch_d'].iloc[-1] < 20 and df['stoch_k'].iloc[-1] > df['stoch_k'].iloc[-2],
            'obv_uptrend': df['obv'].iloc[-1] > df['obv_ema'].iloc[-1],
            'parabolic_sar_long': df['close'].iloc[-1] > df['psar'].iloc[-1],
            
            # Условия для Short
            'price_below_cloud': df['close'].iloc[-1] < df['senkou_span_a'].iloc[-1] and df['close'].iloc[-1] < df['senkou_span_b'].iloc[-1],
            'macd_below_signal': df['macd'].iloc[-1] < df['macd_signal'].iloc[-1],
            'rsi_overbought': df['rsi'].iloc[-1] > 70 and df['rsi'].iloc[-1] < df['rsi'].iloc[-2],
            'price_above_bb_upper': df['close'].iloc[-1] > df['bb_upper'].iloc[-1],
            'ema_crossover_short': df['ema_9'].iloc[-1] < df['ema_21'].iloc[-1] and df['ema_9'].iloc[-2] >= df['ema_21'].iloc[-2],
            'stoch_overbought': df['stoch_k'].iloc[-1] > 80 and df['stoch_d'].iloc[-1] > 80 and df['stoch_k'].iloc[-1] < df['stoch_k'].iloc[-2],
            'obv_downtrend': df['obv'].iloc[-1] < df['obv_ema'].iloc[-1],
            'parabolic_sar_short': df['close'].iloc[-1] < df['psar'].iloc[-1],
            
            # Общие индикаторы
            'adx_strong_trend': df['adx'].iloc[-1] > 25,
            'volume_spike': df['volume'].iloc[-1] > df['volume'].rolling(window=20).mean().iloc[-1] * 1.5,
            'williams_r_bullish': df['wr'].iloc[-1] < -80 and df['wr'].iloc[-1] > df['wr'].iloc[-2],
            'williams_r_bearish': df['wr'].iloc[-1] > -20 and df['wr'].iloc[-1] < df['wr'].iloc[-2],
            
            # Supertrend индикаторы
            'supertrend_value': df['supertrend'].iloc[-1] if len(df) > atr_period else None,
            'supertrend_long': df['close'].iloc[-1] > df['supertrend'].iloc[-1] if len(df) > atr_period else False,
            'supertrend_short': df['close'].iloc[-1] < df['supertrend'].iloc[-1] if len(df) > atr_period else False
        }
        
        # Pattern Recognition Integration
        if config['use_pattern_recognition']:
            try:
                # Analyze patterns using the pattern recognition module
                pattern_results = analyze_patterns(df, config['pattern_lookback_period'])
                
                # Add pattern recognition info to indicators
                indicators['pattern_bullish_score'] = pattern_results['bullish_score']
                indicators['pattern_bearish_score'] = pattern_results['bearish_score']
                indicators['pattern_bullish'] = len(pattern_results['total_patterns']['bullish']) > 0
                indicators['pattern_bearish'] = len(pattern_results['total_patterns']['bearish']) > 0
                
                # Store detailed pattern information
                indicators['candlestick_patterns'] = pattern_results['candlestick_patterns']
                indicators['chart_patterns'] = pattern_results['chart_patterns']
                
                logger.debug(f"Pattern recognition: Bullish={indicators['pattern_bullish_score']}, Bearish={indicators['pattern_bearish_score']}")
            except Exception as e:
                logger.error(f"Error in pattern recognition: {e}")
        
        return indicators
    except Exception as e:
        logger.error(f"Ошибка расчёта индикаторов: {e}")
        return None

# Проверка условий для лонга
def check_long_conditions(ind, df):
    """Проверка условий для Long сигнала"""
    conditions = {
        'price_above_cloud': ind.get('price_above_cloud', False),
        'macd_above_signal': ind.get('macd_above_signal', False),
        'rsi_oversold': ind.get('rsi_oversold', False),
        'price_below_bb_lower': ind.get('price_below_bb_lower', False),
        'ema_crossover_long': ind.get('ema_crossover_long', False),
        'stoch_oversold': ind.get('stoch_oversold', False),
        'obv_uptrend': ind.get('obv_uptrend', False),
        'parabolic_sar_long': ind.get('parabolic_sar_long', False),
        'adx_strong_trend': ind.get('adx_strong_trend', False),
        'volume_spike': ind.get('volume_spike', False),
        'williams_r_bullish': ind.get('williams_r_bullish', False),
        'supertrend_long': ind.get('supertrend_long', False)
    }
    
    # BB+Supertrend комбинация
    bb_width = ind.get('bb_width', 0)
    conditions['bbtrend_supertrend_long'] = (
        ind.get('price_below_bb_lower', False) and 
        ind.get('supertrend_long', False) and 
        bb_width > 0.1
    )
    
    # Add pattern recognition conditions if available
    if config.get('use_pattern_recognition', False):
        if 'pattern_bullish_score' in ind and 'pattern_bearish_score' in ind:
            # Pattern bullish score > bearish score is a bullish signal
            conditions['pattern_bullish'] = ind.get('pattern_bullish_score', 0) > ind.get('pattern_bearish_score', 0)
            
            # Add each individual candlestick pattern if present
            if 'candlestick_patterns' in ind and 'bullish' in ind['candlestick_patterns']:
                for pattern in ind['candlestick_patterns']['bullish']:
                    conditions[f'cdl_{pattern}'] = True
                    
            # Add each individual chart pattern if present
            if 'chart_patterns' in ind and 'bullish' in ind['chart_patterns']:
                for pattern in ind['chart_patterns']['bullish']:
                    conditions[f'chart_{pattern}'] = True
    
    return conditions

# Проверка условий для шорта
def check_short_conditions(ind, df):
    """Проверка условий для Short сигнала"""
    conditions = {
        'price_below_cloud': ind.get('price_below_cloud', False),
        'macd_below_signal': ind.get('macd_below_signal', False),
        'rsi_overbought': ind.get('rsi_overbought', False),
        'price_above_bb_upper': ind.get('price_above_bb_upper', False),
        'ema_crossover_short': ind.get('ema_crossover_short', False),
        'stoch_overbought': ind.get('stoch_overbought', False),
        'obv_downtrend': ind.get('obv_downtrend', False),
        'parabolic_sar_short': ind.get('parabolic_sar_short', False),
        'adx_strong_trend': ind.get('adx_strong_trend', False),
        'volume_spike': ind.get('volume_spike', False),
        'williams_r_bearish': ind.get('williams_r_bearish', False),
        'supertrend_short': ind.get('supertrend_short', False)
    }
    
    # BB+Supertrend комбинация
    bb_width = ind.get('bb_width', 0)
    conditions['bbtrend_supertrend_short'] = (
        ind.get('price_above_bb_upper', False) and 
        ind.get('supertrend_short', False) and 
        bb_width > 0.1
    )
    
    # Add pattern recognition conditions if available
    if config.get('use_pattern_recognition', False):
        if 'pattern_bullish_score' in ind and 'pattern_bearish_score' in ind:
            # Pattern bearish score > bullish score is a bearish signal
            conditions['pattern_bearish'] = ind.get('pattern_bearish_score', 0) > ind.get('pattern_bullish_score', 0)
            
            # Add each individual candlestick pattern if present
            if 'candlestick_patterns' in ind and 'bearish' in ind['candlestick_patterns']:
                for pattern in ind['candlestick_patterns']['bearish']:
                    conditions[f'cdl_{pattern}'] = True
                    
            # Add each individual chart pattern if present
            if 'chart_patterns' in ind and 'bearish' in ind['chart_patterns']:
                for pattern in ind['chart_patterns']['bearish']:
                    conditions[f'chart_{pattern}'] = True
    
    return conditions

# Комбинации высокой уверенности
high_confidence_combinations = [
    ('rsi_oversold', 'macd_above_signal', 'adx_strong_trend'),
    ('price_above_cloud', 'ema_crossover_long', 'volume_spike'),
    ('rsi_oversold', 'price_below_bb_lower', 'stoch_oversold')
]

# Создание сигнала
def create_signal(symbol, signal_type, current_price, conditions, df, volume_usd, news):
    """Создание структуры сигнала"""
    
    # Определение, является ли монета мемной
    base_symbol = symbol.split('/')[0].replace('USDT:USDT', '').replace('USDT', '')
    is_meme_coin = any(meme in base_symbol for meme in config['meme_coins'])
    
    # Настройка параметров в зависимости от типа монеты
    if is_meme_coin:
        leverage = config['meme_leverage']
        tp_percent = config['meme_tp_percent']
        offset_percent = config.get('meme_supertrend_offset_percent', 1.0)
        offset_atr = config.get('meme_supertrend_offset_atr', 1.0)
        sl_atr_multiplier = config.get('meme_sl_atr_multiplier', 2.0)
    else:
        leverage = config['leverage']
        tp_percent = config['tp_percent']
        offset_percent = config.get('supertrend_offset_percent', 0.5)
        offset_atr = config.get('supertrend_offset_atr', 0.5)
        sl_atr_multiplier = config.get('sl_atr_multiplier', 1.5)
    
    # Доступ к ATR и значению суперттренда
    atr_value = float(conditions.get('atr', 0))
    supertrend_value = conditions.get('supertrend_value')
    
    # Инициализация entry_price
    entry_price = current_price
    
    # Проверка доступности суперттренда
    if supertrend_value is None:
        logger.debug(f"{symbol}: Суперттренд не рассчитан, используются стандартные TP/SL")
        # Расчет TP и SL стандартным методом
        sl_percent = config['sl_percent']
        
        if signal_type == 'long':
            tp = round(entry_price * (1 + tp_percent / 100), 6)
            sl = round(entry_price * (1 - sl_percent / 100), 6)
        else:  # short
            tp = round(entry_price * (1 - tp_percent / 100), 6)
            sl = round(entry_price * (1 + sl_percent / 100), 6)
    else:
        # Определение значения отступа (максимальное из процента и ATR)
        offset_by_percent = current_price * offset_percent / 100
        offset_by_atr = atr_value * offset_atr
        offset = max(offset_by_percent, offset_by_atr) if atr_value > 0 else offset_by_percent
        
        logger.debug(f"{symbol}: Отступ для {signal_type}: {offset:.6f} ({offset_percent}% или {offset_atr}*ATR)")
        
        # Настройка TP и SL на основе суперттренда
        if signal_type == 'long':
            # Проверка, соответствует ли цена условиям входа с отступом
            entry_with_offset = supertrend_value + offset
            if current_price < entry_with_offset:
                logger.debug(f"{symbol}: Цена {current_price} ниже точки входа с отступом {entry_with_offset}")
                # Устанавливаем текущую цену как точку входа (без отступа)
                entry_price = current_price
            else:
                # Используем точку входа с отступом
                entry_price = entry_with_offset
                logger.debug(f"{symbol}: Установлена точка входа с отступом: {entry_price}")
            
            # TP установлен выше суперттренда
            tp = round(supertrend_value * 1.05, 6)
            
            # SL на основе ATR (если доступен) или стандартный процентный SL
            if atr_value > 0:
                sl = round(entry_price - (atr_value * sl_atr_multiplier), 6)
            else:
                sl_percent = config['sl_percent']
                sl = round(entry_price * (1 - sl_percent / 100), 6)
        else:  # short
            # Проверка, соответствует ли цена условиям входа с отступом
            entry_with_offset = supertrend_value - offset
            if current_price > entry_with_offset:
                logger.debug(f"{symbol}: Цена {current_price} выше точки входа с отступом {entry_with_offset}")
                # Устанавливаем текущую цену как точку входа (без отступа)
                entry_price = current_price
            else:
                # Используем точку входа с отступом
                entry_price = entry_with_offset
                logger.debug(f"{symbol}: Установлена точка входа с отступом: {entry_price}")
            
            # TP установлен ниже суперттренда
            tp = round(supertrend_value * 0.95, 6)
            
            # SL на основе ATR (если доступен) или стандартный процентный SL
            if atr_value > 0:
                sl = round(entry_price + (atr_value * sl_atr_multiplier), 6)
            else:
                sl_percent = config['sl_percent']
                sl = round(entry_price * (1 + sl_percent / 100), 6)
    
    # Проверка на High Confidence комбинации индикаторов
    high_confidence_combo = []
    
    # Считаем общее количество сработавших индикаторов
    active_indicators = [key for key, value in conditions.items() if value and not key.startswith('cdl_') and not key.startswith('chart_')]
    
    # Проверка на наличие комбинаций высокой уверенности
    for combo in high_confidence_combinations:
        if all(ind in active_indicators for ind in combo):
            high_confidence_combo = combo
            break
    
    # Расчёт процентов прибыли/убытка
    tp_percent_calc = abs((tp - entry_price) / entry_price * 100)
    sl_percent_calc = abs((sl - entry_price) / entry_price * 100)
    
    # Логирование сигнала
    logger.debug(f"Сигнал: {signal_type}, Вход: {entry_price}, TP: {tp}, SL: {sl}")
    logger.debug(f"Активные индикаторы ({len(active_indicators)}): {active_indicators}")
    if high_confidence_combo:
        logger.debug(f"Комбинация высокой уверенности: {high_confidence_combo}")
    
    return {
        'symbol': symbol,
        'type': signal_type,
        'entry_price': entry_price,
        'tp': tp,
        'sl': sl,
        'tp_percent': tp_percent_calc,
        'sl_percent': sl_percent_calc,
        'indicators': conditions,
        'volume_usd': volume_usd,
        'timestamp': datetime.utcnow().isoformat() + 'Z',
        'high_confidence_combo': high_confidence_combo,
        'news': news,
        'leverage': leverage,
        'position_size': config['position_size_usd'],
        'active_indicators': active_indicators,
        'active_indicators_count': len(active_indicators),
        'supertrend_value': supertrend_value
    }

# Отправка уведомления о сигнале
async def send_signal_notification(signal):
    """Отправка оповещения о новом сигнале в Telegram"""
    try:
        symbol = signal['symbol'].replace('USDT:USDT', 'USDT')
        
        # Эмодзи и стиль в зависимости от типа сигнала
        if signal['type'] == 'long':
            emoji = "📈"
            signal_type = "LONG 🟢"
        else:
            emoji = "📉"
            signal_type = "SHORT 🔴"
        
        # Расчёт TP и SL для мемных монет
        base_symbol = symbol.split('/')[0].replace('USDT', '')
        
        if any(meme in base_symbol for meme in config['meme_coins']):
            tp_percent = config['meme_tp_percent']
            leverage = config['meme_leverage']
            meme_tag = " 🚨MEME🚨"
        else:
            tp_percent = config['tp_percent']
            leverage = config['leverage']
            meme_tag = ""
        
        entry_price = signal['entry_price']
        tp_price = signal['tp']
        sl_price = signal['sl']
        
        # Расчёт P/L
        tp_pl_percent = signal.get('tp_percent', tp_percent)
        sl_pl_percent = signal.get('sl_percent', config['sl_percent'])
        tp_pl_usd = round(config['position_size_usd'] * leverage * tp_pl_percent / 100, 2)
        sl_pl_usd = round(config['position_size_usd'] * leverage * sl_pl_percent / 100, 2)
        
        # Форматирование сообщения
        message = f"{emoji} {signal_type} {symbol}{meme_tag}\n\n"
        
        # Информация о входе и Супертренде
        if 'supertrend_value' in signal and signal['supertrend_value'] is not None:
            supertrend_value = signal['supertrend_value']
            message += f"💰 Вход: {entry_price}"
            
            # Добавление информации об отступе от Супертренда
            if signal['type'] == 'long':
                if entry_price > supertrend_value:
                    offset = round((entry_price - supertrend_value) / supertrend_value * 100, 2)
                    message += f" (SuperTrend + {offset}%)\n"
                else:
                    # Если точка входа совпадает с текущей ценой (ниже расчетной с отступом)
                    message += " (Текущая цена)\n"
            else:  # short
                if entry_price < supertrend_value:
                    offset = round((supertrend_value - entry_price) / supertrend_value * 100, 2)
                    message += f" (SuperTrend - {offset}%)\n"
                else:
                    # Если точка входа совпадает с текущей ценой (выше расчетной с отступом)
                    message += " (Текущая цена)\n"
                
            message += f"🎯 TP: {tp_price} (+{tp_pl_percent:.2f}% / +${tp_pl_usd})\n"
            message += f"🛑 SL: {sl_price} (-{sl_pl_percent:.2f}% / -${sl_pl_usd})\n"
            message += f"📊 SuperTrend: {supertrend_value}\n"
        else:
            message += f"💰 Вход: {entry_price}\n"
            message += f"🎯 TP: {tp_price} (+{tp_pl_percent:.2f}% / +${tp_pl_usd})\n"
            message += f"🛑 SL: {sl_price} (-{sl_pl_percent:.2f}% / -${sl_pl_usd})\n"
            
        message += f"⚙️ Леверидж: {leverage}x\n"
        
        # Правильное форматирование объема (в миллионах долларов)
        if 'volume_usd' in signal and signal['volume_usd'] > 0:
            volume_in_millions = round(signal['volume_usd']/1000000, 2)
            message += f"💵 Объём за 24ч: ${volume_in_millions}M\n"
        else:
            message += "💵 Объём за 24ч: N/A\n"
        
        # Fear & Greed Index если доступен
        if 'fng_value' in signal:
            fng_value = signal['fng_value']
            fng_status = "Страх" if fng_value < 30 else "Жадность" if fng_value > 70 else "Нейтрально"
            message += f"😱 Fear & Greed: {fng_value} ({fng_status})\n"
            
            if 'fng_bonus' in signal:
                message += f"⚡ F&G бонус: +{signal['fng_bonus']}\n"
        
        message += "\n"
        
        # ML prediction info if available
        if 'ml_prediction' in signal and signal['ml_prediction']:
            ml_conf = signal['ml_prediction']['confidence'] * 100
            direction_emoji = "📈" if signal['ml_prediction']['direction'] == 'up' else "📉"
            message += f"🤖 ML: {direction_emoji} {ml_conf:.1f}% уверенности"
            
            if 'ml_bonus' in signal:
                message += f" (+{signal['ml_bonus']} к счёту)"
                
            message += "\n\n"
        
        # Ключевые индикаторы
        key_indicators = [
            'price_above_cloud' if signal['type'] == 'long' else 'price_below_cloud',
            'macd_above_signal' if signal['type'] == 'long' else 'macd_below_signal',
            'rsi_oversold' if signal['type'] == 'long' else 'rsi_overbought',
            'price_below_bb_lower' if signal['type'] == 'long' else 'price_above_bb_upper',
            'ema_crossover_long' if signal['type'] == 'long' else 'ema_crossover_short',
            'stoch_oversold' if signal['type'] == 'long' else 'stoch_overbought',
            'obv_uptrend' if signal['type'] == 'long' else 'obv_downtrend',
            'parabolic_sar_long' if signal['type'] == 'long' else 'parabolic_sar_short',
            'adx_strong_trend',
            'volume_spike',
            'supertrend_long' if signal['type'] == 'long' else 'supertrend_short',
            'bbtrend_supertrend_long' if signal['type'] == 'long' else 'bbtrend_supertrend_short'
        ]
        indicator_names = {
            'price_above_cloud': 'Ichimoku Cloud',
            'price_below_cloud': 'Ichimoku Cloud',
            'macd_above_signal': 'MACD Above Signal',
            'macd_below_signal': 'MACD Below Signal',
            'rsi_oversold': 'RSI Oversold',
            'rsi_overbought': 'RSI Overbought',
            'price_below_bb_lower': 'Bollinger Bands Lower',
            'price_above_bb_upper': 'Bollinger Bands Upper',
            'ema_crossover_long': 'EMA Crossover',
            'ema_crossover_short': 'EMA Crossover',
            'stoch_oversold': 'Stochastic Oversold',
            'stoch_overbought': 'Stochastic Overbought',
            'obv_uptrend': 'OBV Uptrend',
            'obv_downtrend': 'OBV Downtrend',
            'parabolic_sar_long': 'Parabolic SAR',
            'parabolic_sar_short': 'Parabolic SAR',
            'adx_strong_trend': 'ADX Strong Trend',
            'volume_spike': 'Volume Spike',
            'supertrend_long': 'SuperTrend Long',
            'supertrend_short': 'SuperTrend Short',
            'bbtrend_supertrend_long': 'BB+SuperTrend Long',
            'bbtrend_supertrend_short': 'BB+SuperTrend Short'
        }
        
        # Подсчитываем активные индикаторы (исключая паттерны cdl_ и chart_)
        active_count = sum(
            1 for key, value in signal['indicators'].items() 
            if value and not key.startswith('cdl_') and not key.startswith('chart_')
        )
        
        total_indicators = len([
            key for key in signal['indicators'].keys() 
            if not key.startswith('cdl_') and not key.startswith('chart_')
        ])
        
        message += f"🔍 Индикаторы [{active_count}/{total_indicators}]:\n"
        for ind in key_indicators:
            status = "✔️" if signal['indicators'].get(ind) else "✘"
            message += f"  [{status}] {indicator_names.get(ind, ind)}\n"
        
        # Выделение стратегии BBTrend+SuperTrend
        if signal['indicators'].get('bbtrend_supertrend_long' if signal['type'] == 'long' else 'bbtrend_supertrend_short', False):
            message += f"\n✨ Стратегия BB+SuperTrend активна!\n"
        
        if signal['high_confidence_combo']:
            message += f"🔥 High Confidence: {', '.join(indicator_names.get(ind, ind) for ind in signal['high_confidence_combo'])}\n"
        
        message += f"📰 Новости: {signal['news']}\n"
        message += f"⚠️ Не финансовая рекомендация!"
        
        await send_telegram_message(message)
        logger.info(f"Сигнал отправлен: {signal['type']} на {signal['symbol']}")
    except Exception as e:
        logger.error(f"Ошибка отправки сигнала: {e}")

# Глобальный словарь для хранения последних сигналов
last_signals = {}

# Анализ сигналов
async def analyze_signals(symbol, indicators, df, volume_usd, price_change_1h):
    try:
        if indicators is None:
            logger.debug(f"Пропущен {symbol}: индикаторы не рассчитаны")
            return None
        
        long_conditions = check_long_conditions(indicators, df)
        short_conditions = check_short_conditions(indicators, df)
        
        # Count normal indicators
        long_score = sum(
            value for key, value in long_conditions.items() 
            if not (key.startswith('cdl_') or key.startswith('chart_') or key == 'pattern_bullish')
        )
        short_score = sum(
            value for key, value in short_conditions.items() 
            if not (key.startswith('cdl_') or key.startswith('chart_') or key == 'pattern_bearish')
        )
        
        # Логирование активных индикаторов
        active_long_indicators = [key for key, value in long_conditions.items() if value and not (key.startswith('cdl_') or key.startswith('chart_'))]
        active_short_indicators = [key for key, value in short_conditions.items() if value and not (key.startswith('cdl_') or key.startswith('chart_'))]
        
        logger.debug(f"Активные long индикаторы для {symbol}: {active_long_indicators}")
        logger.debug(f"Активные short индикаторы для {symbol}: {active_short_indicators}")
        
        # Проверка стратегии BBTrend+SuperTrend
        bbtrend_supertrend_long = long_conditions.get('bbtrend_supertrend_long', False)
        bbtrend_supertrend_short = short_conditions.get('bbtrend_supertrend_short', False)
        
        if bbtrend_supertrend_long:
            logger.debug(f"{symbol}: Обнаружена стратегия BB+SuperTrend Long")
        elif bbtrend_supertrend_short:
            logger.debug(f"{symbol}: Обнаружена стратегия BB+SuperTrend Short")
        
        # Интеграция с Fear & Greed Index
        fng_value = indicators.get('fng_value', 50)  # Значение по умолчанию 50 (нейтральное)
        fng_bonus = 0
        
        # Если FNG < 30 (страх) и есть суперттренд лонг - это хорошая точка входа
        if fng_value < 30 and long_conditions.get('supertrend_long', False):
            fng_bonus = 1.5
            long_score += fng_bonus
            logger.debug(f"{symbol}: Fear & Greed: {fng_value} (страх) - бонус к long сигналу: +{fng_bonus}")
            
        # Если FNG > 70 (жадность) и есть суперттренд шорт - это хорошая точка входа
        elif fng_value > 70 and short_conditions.get('supertrend_short', False):
            fng_bonus = 1.5
            short_score += fng_bonus
            logger.debug(f"{symbol}: Fear & Greed: {fng_value} (жадность) - бонус к short сигналу: +{fng_bonus}")
        
        # Add weighted pattern scores if enabled
        pattern_score_long = 0
        pattern_score_short = 0
        
        if config.get('use_pattern_recognition', False):
            # Count candlestick patterns
            cdl_patterns_long = sum(
                1 for key, value in long_conditions.items() 
                if key.startswith('cdl_') and value
            )
            cdl_patterns_short = sum(
                1 for key, value in short_conditions.items() 
                if key.startswith('cdl_') and value
            )
            
            # Count chart patterns
            chart_patterns_long = sum(
                1 for key, value in long_conditions.items() 
                if key.startswith('chart_') and value
            )
            chart_patterns_short = sum(
                1 for key, value in short_conditions.items() 
                if key.startswith('chart_') and value
            )
            
            # Add general pattern match if present
            if long_conditions.get('pattern_bullish', False):
                cdl_patterns_long += 1
            
            if short_conditions.get('pattern_bearish', False):
                cdl_patterns_short += 1
            
            # Add weighted pattern scores
            pattern_weight = config.get('pattern_weight', 2.0)
            pattern_score_long = (cdl_patterns_long + chart_patterns_long) * pattern_weight
            pattern_score_short = (cdl_patterns_short + chart_patterns_short) * pattern_weight
            
            long_score += pattern_score_long
            short_score += pattern_score_short
            
            logger.debug(f"{symbol}: Added pattern weights - Long: +{pattern_score_long}, Short: +{pattern_score_short}")
        
        # Get total count of indicators (excluding patterns which are already weighted)
        total_indicators = len([
            key for key in set(list(long_conditions.keys()) + list(short_conditions.keys()))
            if not (key.startswith('cdl_') or key.startswith('chart_') or key in ['pattern_bullish', 'pattern_bearish'])
        ])
        
        logger.debug(f"{symbol}: Long Score = {long_score}/{total_indicators} (Базовые: {long_score - pattern_score_long - fng_bonus if fng_bonus > 0 and long_score > short_score else long_score - pattern_score_long})")
        logger.debug(f"{symbol}: Short Score = {short_score}/{total_indicators} (Базовые: {short_score - pattern_score_short - fng_bonus if fng_bonus > 0 and short_score > long_score else short_score - pattern_score_short})")
        
        # Get news sentiment
        news = await check_news(symbol, price_change_1h)
        
        # Get ML model prediction if available
        ml_prediction = None
        ml_bonus_long = 0
        ml_bonus_short = 0
        
        if config.get('use_ml_predictions', False):
            ml_prediction = await get_ml_prediction(symbol, ohlcv_to_dataframe(df))
            
            if ml_prediction:
                ml_direction = ml_prediction['direction']
                ml_confidence = ml_prediction['confidence']
                logger.debug(f"ML-прогноз для {symbol}: {ml_direction.upper()} с уверенностью {ml_confidence:.2f}")
                
                # Add ML prediction weight to appropriate score
                if ml_direction == 'up' and ml_confidence >= config.get('ml_confidence_threshold', 0.65):
                    # ML predicts price will go up, add to long score
                    ml_bonus_long = 2
                    long_score += ml_bonus_long
                    logger.debug(f"Добавление ML-веса к long сигналу для {symbol} (+{ml_bonus_long})")
                elif ml_direction == 'down' and ml_confidence >= config.get('ml_confidence_threshold', 0.65):
                    # ML predicts price will go down, add to short score
                    ml_bonus_short = 2
                    short_score += ml_bonus_short
                    logger.debug(f"Добавление ML-веса к short сигналу для {symbol} (+{ml_bonus_short})")
        
        # Фиксированный порог в 8 индикаторов для обычных сигналов, 5 для BB+SuperTrend
        fixed_threshold = 8
        bbtrend_supertrend_threshold = 5
        
        # Создаем сигнал, если число индикаторов выше порога
        signal = None
        
        # Приоритет сигналов на основе комбинации BB + Supertrend
        if bbtrend_supertrend_long and long_score >= bbtrend_supertrend_threshold:
            logger.debug(f"{symbol}: Сигнал Long на основе BB+SuperTrend (скор: {long_score}, основные: {len(active_long_indicators)})")
            signal = create_signal(symbol, 'long', indicators['current_price'], long_conditions, df, volume_usd, news)
            if ml_prediction and ml_prediction['direction'] == 'up':
                signal['ml_prediction'] = {'direction': 'up', 'confidence': ml_prediction['confidence']}
                signal['ml_bonus'] = ml_bonus_long
            
            # Add pattern information to signal if available
            if config['use_pattern_recognition'] and 'candlestick_patterns' in indicators and 'chart_patterns' in indicators:
                signal['patterns'] = {
                    'candlestick': indicators['candlestick_patterns']['bullish'],
                    'chart': indicators['chart_patterns']['bullish']
                }
        elif bbtrend_supertrend_short and short_score >= bbtrend_supertrend_threshold:
            logger.debug(f"{symbol}: Сигнал Short на основе BB+SuperTrend (скор: {short_score}, основные: {len(active_short_indicators)})")
            signal = create_signal(symbol, 'short', indicators['current_price'], short_conditions, df, volume_usd, news)
            if ml_prediction and ml_prediction['direction'] == 'down':
                signal['ml_prediction'] = {'direction': 'down', 'confidence': ml_prediction['confidence']}
                signal['ml_bonus'] = ml_bonus_short
            
            # Add pattern information to signal if available
            if config['use_pattern_recognition'] and 'candlestick_patterns' in indicators and 'chart_patterns' in indicators:
                signal['patterns'] = {
                    'candlestick': indicators['candlestick_patterns']['bearish'],
                    'chart': indicators['chart_patterns']['bearish']
                }
        # Обычные сигналы, если BBTrend не сработал
        elif long_score >= fixed_threshold:
            logger.debug(f"{symbol}: Стандартный Long сигнал (скор: {long_score}, основные: {len(active_long_indicators)})")
            signal = create_signal(symbol, 'long', indicators['current_price'], long_conditions, df, volume_usd, news)
            if ml_prediction and ml_prediction['direction'] == 'up':
                signal['ml_prediction'] = {'direction': 'up', 'confidence': ml_prediction['confidence']}
                signal['ml_bonus'] = ml_bonus_long
            
            # Add pattern information to signal if available
            if config['use_pattern_recognition'] and 'candlestick_patterns' in indicators and 'chart_patterns' in indicators:
                signal['patterns'] = {
                    'candlestick': indicators['candlestick_patterns']['bullish'],
                    'chart': indicators['chart_patterns']['bullish']
                }
        elif short_score >= fixed_threshold:
            logger.debug(f"{symbol}: Стандартный Short сигнал (скор: {short_score}, основные: {len(active_short_indicators)})")
            signal = create_signal(symbol, 'short', indicators['current_price'], short_conditions, df, volume_usd, news)
            if ml_prediction and ml_prediction['direction'] == 'down':
                signal['ml_prediction'] = {'direction': 'down', 'confidence': ml_prediction['confidence']}
                signal['ml_bonus'] = ml_bonus_short
            
            # Add pattern information to signal if available
            if config['use_pattern_recognition'] and 'candlestick_patterns' in indicators and 'chart_patterns' in indicators:
                signal['patterns'] = {
                    'candlestick': indicators['candlestick_patterns']['bearish'],
                    'chart': indicators['chart_patterns']['bearish']
                }
        
        # Проверка повторных сигналов
        if signal:
            global last_signals
            last_signal = last_signals.get(symbol)
            
            if last_signal:
                # Проверка изменения типа сигнала
                if last_signal['type'] == signal['type']:
                    # Если тип сигнала тот же, проверяем изменение цены входа
                    price_diff = abs(signal['entry_price'] - last_signal['entry_price']) / last_signal['entry_price'] * 100
                    if price_diff < 1.0:
                        logger.debug(f"Пропуск повторного сигнала для {symbol}: тип {signal['type']}, изменение цены всего {price_diff:.2f}%")
                        return None
                    else:
                        logger.debug(f"Новый сигнал для {symbol} с тем же типом, но с изменением цены {price_diff:.2f}%")
                else:
                    logger.debug(f"Новый сигнал для {symbol} с измененным типом: {last_signal['type']} -> {signal['type']}")
            
            # Добавляем информацию о F&G бонусе в сигнал
            if fng_bonus > 0:
                signal['fng_bonus'] = fng_bonus
                signal['fng_value'] = fng_value
            
            # Сохраняем сигнал как последний для этого символа
            last_signals[symbol] = signal
            
            return signal
        
        return None
    except Exception as e:
        logger.error(f"Ошибка анализа сигналов для {symbol}: {e}")
        return None

# Обработка пары
async def process_pair(symbol, timeframe=config['timeframe'], limit=config['ohlcv_limit'], volume_data=None):
    logger.debug(f"Обработка {symbol}")
    try:
        ohlcv = await load_ohlcv(symbol, timeframe, limit)
        if not ohlcv or len(ohlcv) < limit:
            logger.debug(f"Пропущен {symbol}: недостаточно данных OHLCV (получено {len(ohlcv) if ohlcv else 0} строк)")
            return None

        df = pd.DataFrame(ohlcv, columns=['timestamp', 'open', 'high', 'low', 'close', 'volume'])
        df['timestamp'] = pd.to_datetime(df['timestamp'], unit='ms')
        
        # Проверка на NaN и минимальное количество данных
        if df[['open', 'high', 'low', 'close', 'volume']].isna().any().any() or len(df) < 14:
            logger.debug(f"Пропущен {symbol}: NaN в данных или недостаточно строк ({len(df)})")
            return None

        # Проверка пампа
        current_price = df['close'].iloc[-1]
        price_change_1h = abs((current_price - df['close'].iloc[-2]) / df['close'].iloc[-2] * 100) if len(df) > 1 and df['close'].iloc[-2] != 0 else 0
        if price_change_1h > config['pump_price_threshold']:
            logger.debug(f"Пропущен {symbol}: памп (рост {price_change_1h:.2f}%)")
            return None

        # Проверка объёма
        avg_volume = df['volume'].rolling(window=20).mean().iloc[-1]
        if not pd.isna(avg_volume) and df['volume'].iloc[-1] != 0:
            volume_spike = df['volume'].iloc[-1] > avg_volume * config['pump_volume_multiplier']
            if volume_spike:
                logger.debug(f"Пропущен {symbol}: Volume Spike (x{config['pump_volume_multiplier']})")
                return None

        # Получение Fear & Greed
        fng_value = await fetch_fear_and_greed()
        
        # Расчёт индикаторов
        indicators = calculate_indicators(df, fng_value)
        if indicators is None:
            logger.debug(f"Пропущен {symbol}: не удалось рассчитать индикаторы")
            return None
        
        # Проверка волатильности
        if pd.isna(indicators['atr']) or pd.isna(indicators['avg_atr']) or indicators['atr'] < indicators['avg_atr'] * config['atr_volatility_multiplier']:
            logger.debug(f"Пропущен {symbol}: низкая волатильность или NaN в ATR")
            return None
            
        # Предварительный запуск ML-модели для обучения в фоне, если ещё не обучена
        if config['use_ml_predictions']:
            model_dir = 'models'
            model_path = os.path.join(model_dir, f'{symbol.replace("/", "_")}_{timeframe}_model.pkl')
            if not os.path.exists(model_path):
                asyncio.create_task(train_ml_model(symbol, timeframe))
                logger.debug(f"Запущено фоновое обучение ML-модели для {symbol}")

        # Анализ сигналов
        volume_usd = volume_data.get(symbol, 0)
        signal = await analyze_signals(symbol, indicators, df, volume_usd, price_change_1h)
        if signal:
            await send_signal_notification(signal)
            log_signal_to_csv(signal)
        
        return signal
    except Exception as e:
        logger.error(f"Ошибка обработки {symbol}: {e}")
        return None

# Бэктест символа
async def backtest_symbol(symbol, timeframe='1h', start_date=None, end_date=None, send_notifications=False):
    """Тестирование сигналов на исторических данных."""
    logger.info(f"Запуск бэктеста для {symbol}")
    try:
        # Конвертация дат в миллисекунды
        since = int(pd.Timestamp(start_date).timestamp() * 1000) if start_date else None
        ohlcv = await bybit.fetch_ohlcv(symbol, timeframe, since=since, limit=1000)
        if not ohlcv or len(ohlcv) < 20:
            logger.warning(f"Недостаточно данных для {symbol}: {len(ohlcv) if ohlcv else 0} строк")
            return None

        df = pd.DataFrame(ohlcv, columns=['timestamp', 'open', 'high', 'low', 'close', 'volume'])
        df['timestamp'] = pd.to_datetime(df['timestamp'], unit='ms')
        
        signals = []
        fng_value = await fetch_fear_and_greed()
        
        # Симуляция обработки для каждого бара
        for i in range(20, len(df)):
            window_df = df.iloc[:i].copy()
            indicators = calculate_indicators(window_df, fng_value)
            if indicators is None:
                continue
            
            price_change_1h = abs((window_df['close'].iloc[-1] - window_df['close'].iloc[-2]) / window_df['close'].iloc[-2] * 100) if len(window_df) > 1 else 0
            volume_usd = window_df['volume'].iloc[-1] * window_df['close'].iloc[-1]
            
            signal = await analyze_signals(symbol, indicators, window_df, volume_usd, price_change_1h)
            if signal:
                signals.append(signal)
                logger.debug(f"Бэктест: Найден сигнал для {symbol} на {window_df['timestamp'].iloc[-1]}")
                if send_notifications:
                    await send_signal_notification(signal)
                log_signal_to_csv(signal)
        
        logger.info(f"Бэктест завершён: {len(signals)} сигналов для {symbol}")
        return signals
    except Exception as e:
        logger.error(f"Ошибка бэктеста для {symbol}: {e}")
        return None

# Функции для ML-моделей
async def train_ml_model(symbol, timeframe=config['timeframe'], force_retrain=False):
    """Обучение ML-модели для прогнозирования тренда"""
    try:
        model_dir = 'models'
        os.makedirs(model_dir, exist_ok=True)
        
        model = CryptoMLPredictor(symbol, timeframe, model_dir)
        model_exists = model.load_model()
        
        # Проверяем, нужно ли переобучать модель
        if model_exists and not force_retrain:
            model_path = model.model_path
            if os.path.exists(model_path):
                mtime = datetime.fromtimestamp(os.path.getmtime(model_path))
                days_since_train = (datetime.now() - mtime).days
                if days_since_train < config['ml_retrain_days']:
                    logger.debug(f"Модель для {symbol} уже обучена и актуальна ({days_since_train} дней назад)")
                    return True
        
        # Загружаем исторические данные (больше свечей для обучения)
        ohlcv = await load_ohlcv(symbol, timeframe, limit=240)  # Больше данных для обучения
        if not ohlcv or len(ohlcv) < 100:
            logger.warning(f"Недостаточно данных для обучения ML-модели {symbol}: {len(ohlcv) if ohlcv else 0} свечей")
            return False
        
        # Конвертируем в DataFrame
        df = ohlcv_to_dataframe(ohlcv)
        
        # Обучаем модель
        result = model.train(df)
        accuracy = result['test_accuracy']
        logger.info(f"ML-модель для {symbol} обучена с точностью {accuracy:.2f}")
        
        return accuracy >= config['min_ml_accuracy']
    
    except Exception as e:
        logger.error(f"Ошибка обучения ML-модели для {symbol}: {e}")
        return False

async def get_ml_prediction(symbol, df=None, timeframe=config['timeframe']):
    """Получение прогноза от ML-модели"""
    try:
        if not config['use_ml_predictions']:
            return None
            
        model = CryptoMLPredictor(symbol, timeframe, 'models')
        
        # Загружаем модель, если она существует
        if not model.load_model():
            # Обучаем новую модель, если нет
            success = await train_ml_model(symbol, timeframe)
            if not success:
                return None
        
        # Если DataFrame не предоставлен, загружаем данные
        if df is None:
            ohlcv = await load_ohlcv(symbol, timeframe)
            if not ohlcv:
                return None
            df = ohlcv_to_dataframe(ohlcv)
        
        # Получаем прогноз
        prediction = model.predict(df)
        
        if not prediction:
            return None
            
        # Проверяем уверенность прогноза
        if prediction['confidence'] < config['ml_confidence_threshold']:
            logger.debug(f"ML-прогноз для {symbol} имеет слишком низкую уверенность: {prediction['confidence']:.2f}")
            return None
            
        logger.info(f"ML-прогноз для {symbol}: {prediction['direction'].upper()} с уверенностью {prediction['confidence']:.2f}")
        
        return prediction
        
    except Exception as e:
        logger.error(f"Ошибка получения ML-прогноза для {symbol}: {e}")
        return None

async def batch_train_ml_models(symbols, timeframe=config['timeframe']):
    """Пакетное обучение ML-моделей для списка символов"""
    if not config['use_ml_predictions']:
        return
    
    logger.info(f"Начало пакетного обучения ML-моделей для {len(symbols)} символов")
    
    # Создаем директорию для моделей, если не существует
    os.makedirs('models', exist_ok=True)
    
    # Загружаем данные для всех символов
    symbol_data = {}
    for symbol in symbols:
        try:
            ohlcv = await load_ohlcv(symbol, timeframe, limit=240)  # Больше данных для обучения
            if ohlcv and len(ohlcv) >= 100:
                df = ohlcv_to_dataframe(ohlcv)
                symbol_data[symbol] = df
            else:
                logger.warning(f"Недостаточно данных для обучения ML-модели {symbol}")
        except Exception as e:
            logger.error(f"Ошибка загрузки данных для ML-модели {symbol}: {e}")
    
    if not symbol_data:
        logger.warning("Нет данных для обучения ML-моделей")
        return
    
    # Обучаем модели пакетно
    results = await batch_train_models(symbol_data, timeframe, 'models')
    
    # Выводим результаты
    successful = sum(1 for result in results.values() if result.get('success', False))
    logger.info(f"Завершено обучение ML-моделей: {successful}/{len(results)} успешно")
    
    # Возвращаем символы с хорошими моделями
    good_models = [
        symbol for symbol, result in results.items()
        if result.get('success', False) and result.get('accuracy', 0) >= config['min_ml_accuracy']
    ]
    
    return good_models

# Основной цикл
async def main():
    try:
        # Тестируем подключение к Telegram при запуске
        await test_telegram()
        
        # Очистка старого кэша
        clean_old_cache()
        
        # Загрузка рынков
        symbols = await load_markets()
        if not symbols:
            logger.error("Не удалось загрузить рынки")
            return
        
        # Проверка аргументов
        parser = argparse.ArgumentParser(description='Crypto Signals Bot')
        parser.add_argument('--backtest', help='Run backtest for symbol', type=str, default=None)
        parser.add_argument('--start-date', help='Start date for backtest (YYYY-MM-DD)', type=str, default=None)
        parser.add_argument('--train-ml', help='Train ML models for all high-volume symbols', action='store_true')
        parser.add_argument('--generate-ml-report', help='Generate ML models analytics report', action='store_true')
        args = parser.parse_args()
        
        # Инициализация
        clean_old_cache()
        
        # Отправка тестового сообщения
        await test_telegram()
        
        # Проверка аргументов
        if args.backtest:
            if args.backtest == 'all':
                # Backtest all high volume pairs
                symbols = await load_markets()
                high_volume_symbols, _ = await filter_high_volume_symbols(symbols)
                
                for symbol in high_volume_symbols[:10]:  # limit to 10 symbols for demo
                    signals = await backtest_symbol(symbol, start_date=args.start_date)
                    if signals:
                        logger.info(f"Backtest {symbol}: {len(signals)} signals generated")
                    await asyncio.sleep(1)
            else:
                # Backtest specific symbol
                signals = await backtest_symbol(args.backtest, start_date=args.start_date)
                if signals:
                    logger.info(f"Backtest {args.backtest}: {len(signals)} signals generated")
            
            # Generate analytics after backtest
            generate_analytics()
            return
        
        # Train ML models if requested
        if args.train_ml:
            symbols = await load_markets()
            high_volume_symbols, _ = await filter_high_volume_symbols(symbols)
            logger.info(f"Starting batch ML training for {len(high_volume_symbols)} high-volume symbols")
            good_models = await batch_train_ml_models(high_volume_symbols)
            logger.info(f"Trained {len(good_models) if good_models else 0} good ML models")
            return
        
        # Generate ML report if requested
        if args.generate_ml_report:
            generate_ml_analytics_report()
            return
        
        # Основной цикл
        while True:
            try:
                # Загрузка доступных рынков
                symbols = await load_markets()
                
                if not symbols:
                    logger.error("Не удалось загрузить символы. Повторная попытка через 5 минут.")
                    await asyncio.sleep(300)
                    continue
                
                # Фильтр по объёму
                high_volume_symbols, volume_data = await filter_high_volume_symbols(symbols)
                
                logger.info(f"Обработка {len(high_volume_symbols)} пар с высоким объёмом")
                
                # Обучение ML-моделей для всех пар с высоким объёмом (в фоне)
                if config['use_ml_predictions']:
                    asyncio.create_task(batch_train_ml_models(high_volume_symbols))
                    logger.info("Запущено фоновое обучение ML-моделей")
                
                # Обработка пар пачками
                signals = []
                chunk_size = 10
                
                for i in range(0, len(high_volume_symbols), chunk_size):
                    chunk = high_volume_symbols[i:i + chunk_size]
                    logger.info(f"Обработка пачки {i//chunk_size + 1}/{len(high_volume_symbols) // chunk_size + 1}: {len(chunk)} пар")
                    
                    # Асинхронная обработка пачки
                    tasks = [process_pair(symbol, volume_data=volume_data) for symbol in chunk]
                    results = await asyncio.gather(*tasks, return_exceptions=True)
                    
                    # Фильтрация ошибок и None значений
                    chunk_signals = [r for r in results if r is not None and not isinstance(r, Exception)]
                    signals.extend(chunk_signals)
                    
                    logger.info(f"Найдено сигналов в текущей пачке: {len(chunk_signals)}")
                    await asyncio.sleep(5)  # Пауза между пачками
                
                logger.info(f"Всего найдено сигналов: {len(signals)}")
                
                # Генерация отчета по ML-моделям раз в день
                if datetime.now().hour == 0 and datetime.now().minute < 15:
                    generate_ml_analytics_report()
                
                # Пауза перед следующим циклом
                logger.info(f"Ожидание {60*5} секунд перед следующим сканированием...")
                await asyncio.sleep(60 * 5)  # 5 минут между циклами
                
            except Exception as e:
                logger.error(f"Ошибка в основном цикле: {e}")
                await asyncio.sleep(60)

    except Exception as e:
        logger.error(f"Критическая ошибка: {e}")

async def update_excel():
    """Обновление Excel-файла с историей сигналов"""
    try:
        export_csv_to_excel()
        logger.debug("Excel-файл обновлен")
    except Exception as e:
        logger.error(f"Ошибка обновления Excel: {e}")

def generate_ml_analytics_report():
    """Генерация отчета по эффективности ML-моделей"""
    try:
        # Check if we have any models trained
        model_dir = 'models'
        os.makedirs(model_dir, exist_ok=True)
        
        if not os.listdir(model_dir):
            logger.warning("Нет обученных ML-моделей для анализа")
            return
            
        # Initialize workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ML Models Analysis"
        
        # Set headers
        headers = [
            "Symbol", "Model Type", "Accuracy", "Top Features", 
            "Signal Win Rate", "Total Signals", "Long Accuracy", "Short Accuracy",
            "Avg Confidence", "Last Trained"
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
            
        # Load signals from CSV to calculate win rates
        signals_df = None
        if os.path.exists('signals_history.csv'):
            try:
                signals_df = pd.read_csv('signals_history.csv')
                # Convert to proper types and handle missing values
                signals_df = signals_df.fillna('')
            except Exception as e:
                logger.error(f"Ошибка загрузки signals_history.csv: {e}")
                
        # Process each model
        row = 2
        for filename in os.listdir(model_dir):
            if filename.endswith('_model.pkl'):
                try:
                    # Parse model info from filename
                    parts = filename.replace('_model.pkl', '').split('_')
                    symbol = parts[0]
                    timeframe = parts[1] if len(parts) > 1 else '1h'
                    
                    # Load model and feature importance
                    model_path = os.path.join(model_dir, filename)
                    feature_path = os.path.join(model_dir, f"{symbol}_{timeframe}_feature_importance.pkl")
                    
                    if not os.path.exists(feature_path):
                        continue
                        
                    model = joblib.load(model_path)
                    feature_importance = joblib.load(feature_path)
                    
                    # Get top 3 features
                    top_features = sorted(
                        feature_importance.items(),
                        key=lambda x: x[1],
                        reverse=True
                    )[:3]
                    top_features_str = ", ".join([f"{f[0]} ({f[1]:.3f})" for f in top_features])
                    
                    # Get model last trained date
                    last_trained = datetime.fromtimestamp(os.path.getmtime(model_path)).strftime('%Y-%m-%d %H:%M')
                    
                    # Calculate signal accuracy using ML predictions
                    ml_accuracy = None
                    signals_count = 0
                    long_accuracy = None
                    short_accuracy = None
                    avg_confidence = None
                    
                    if signals_df is not None and 'ml_direction' in signals_df.columns:
                        # Filter signals with this symbol and ML predictions
                        symbol_signals = signals_df[signals_df['symbol'] == symbol]
                        symbol_signals = symbol_signals[symbol_signals['ml_direction'] != 'none']
                        
                        if len(symbol_signals) > 0:
                            signals_count = len(symbol_signals)
                            
                            # Count successful signals (those that hit TP)
                            successful = symbol_signals[symbol_signals['status'] == 'tp']
                            if len(successful) > 0:
                                ml_accuracy = len(successful) / signals_count
                                
                            # Long accuracy
                            long_signals = symbol_signals[symbol_signals['type'] == 'long']
                            if len(long_signals) > 0:
                                long_successful = long_signals[long_signals['status'] == 'tp']
                                long_accuracy = len(long_successful) / len(long_signals) if len(long_signals) > 0 else None
                                
                            # Short accuracy
                            short_signals = symbol_signals[symbol_signals['type'] == 'short']
                            if len(short_signals) > 0:
                                short_successful = short_signals[short_signals['status'] == 'tp']
                                short_accuracy = len(short_successful) / len(short_signals) if len(short_signals) > 0 else None
                                
                            # Average confidence
                            avg_confidence = symbol_signals['ml_confidence'].astype(float).mean() if 'ml_confidence' in symbol_signals.columns else None
                    
                    # Write to worksheet
                    ws.cell(row=row, column=1, value=symbol)
                    ws.cell(row=row, column=2, value="RandomForest")
                    
                    # Model accuracy (from cross-validation)
                    model_accuracy = getattr(model, 'oob_score_', None)
                    if model_accuracy is None and hasattr(model, 'oob_score'):
                        model_accuracy = model.oob_score
                    ws.cell(row=row, column=3, value=model_accuracy if model_accuracy is not None else "N/A")
                    
                    ws.cell(row=row, column=4, value=top_features_str)
                    ws.cell(row=row, column=5, value=ml_accuracy if ml_accuracy is not None else "N/A")
                    ws.cell(row=row, column=6, value=signals_count)
                    ws.cell(row=row, column=7, value=long_accuracy if long_accuracy is not None else "N/A")
                    ws.cell(row=row, column=8, value=short_accuracy if short_accuracy is not None else "N/A")
                    ws.cell(row=row, column=9, value=avg_confidence if avg_confidence is not None else "N/A")
                    ws.cell(row=row, column=10, value=last_trained)
                    
                    row += 1
                except Exception as e:
                    logger.error(f"Ошибка анализа модели {filename}: {e}")
                    
        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
            
        # Save workbook
        wb.save("ml_analytics_report.xlsx")
        logger.info("Отчет по ML-моделям сгенерирован: ml_analytics_report.xlsx")
        
    except Exception as e:
        logger.error(f"Ошибка генерации отчета по ML-моделям: {e}")

# Запуск основного цикла
if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        logger.info("Бот остановлен пользователем")
    except Exception as e:
        logger.error(f"Критическая ошибка: {e}")